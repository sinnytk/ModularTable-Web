from openpyxl import load_workbook
from datetime import time as Time


class TimeSlot():
    def __init__(self, startTime, endTime):
        self.startTime = startTime
        self.endTime = endTime
    def __repr__(self):
        return "TimeSlot(%s,%s)" % (self.startTime, self.endTime)
class Teacher():
    def __init__(self, teacherName):
        self.teacherName = teacherName
    def __repr__(self):
        return "Teacher(%s)" % (self.teacherName)
    def __eq__(self,other):
        return self.teacherName == other.teacherName
    def __hash__(self):
        return hash(self.__repr__())
class Slot():
    def __init__(self,dayNum,venueNum,timeslotNum,courseNum, teacherNum, sectionNum):
        self.courseNum = courseNum
        self.teacherNum = teacherNum
        self.sectionNum = sectionNum
        self.dayNum = dayNum
        self.venueNum = venueNum
        self.timeslotNum = timeslotNum
    def __repr__(self):
        return "Slot(%s,%s,%s,%s,%s,%s)" % (self.dayNum,self.venueNum,self.timeslotNum, self.courseNum,self.teacherNum,self.sectionNum)
    def __eq__(self, other):
        return (self.courseNum == other.courseNum) and (self.sectionNum == other.sectionNum) and (self.teacherNum == other.teacherNum)
    def __hash__(self):
        return hash(self.__repr__())
class Section():
    def __init__(self,semester,section):
        self.semester = semester
        self.section = section
    def __repr__(self):
        return "Section(%s,%s)" % (self.semester, self.section)
    def __eq__(self, other):
        return (self.semester == other.semester and self.section == other.section)
    def __hash__(self):
        return hash(self.__repr__())
class Course():
    def __init__(self,courseCode):
        self.courseCode = courseCode
    def __repr__(self):
        return "Course(%s)" % (self.courseCode)
    def __eq__(self, other):
        return (self.courseCode == other.courseCode)
    def __hash__(self):
        return hash(self.__repr__())
class Venue():
    def __init__(self, venueName):
        self.venueName = venueName
    def __repr__(self):
        return "Venue(%s)" % (self.venueName)
    def __eq__(self,other):
        return self.venueName == other.venueName
    def __hash__(self):
        return hash(self.__repr__())



def main():
    #dict for records which may be repeated
    #list for records which are constant
    all_venues = {'':None}
    all_teachers = {'':None}
    all_timeslots = {}
    all_courses = {'':None}
    slots = []
    all_sections = {'':None}
    tt = load_workbook('timetable.xlsx')
    sheet = tt.worksheets[0]
    #counts to keep track of new entries and IDs of records
    venue_count = 0
    teacher_count = 0
    timeslot_count = 0
    course_count = 0
    section_count = 0
    for sheetnum, sheet in zip(range(len(tt.worksheets)),tt.worksheets):
        #initializing local venues and timeslots which will be overridden for each sheet(day)
        venues = []
        timeslots = [] 
        # read timeslots
        for cell in sheet[4][2:]:
            if(cell.value == None):
                break
            else:
                time = cell.value.split(' - ')
                start = time[0].split(':')
                end = time[1].split(':')
                timeslots.append(TimeSlot(Time(int(start[0].strip()), int(
                    start[1].strip())), Time(int(end[0].strip()), int(end[1].strip()))))
        # read venues
        for col in sheet.iter_cols(2, 2):
            empty = False
            for cell in col[4:]:
                if(cell.value is not None):
                    venues.append(cell.value.strip().upper())
                elif(empty):
                    break
                else:
                    empty = True
        #saving newly found venues and timeslots in global dictionaries
        for venue in venues:
        if venue not in all_venues:
            venue_count+=1
            all_venues[venue] = venue_count
        for timeslot in timeslots:
        if timeslot not in all_timeslots:
            timeslot_count+=1
            all_timeslots[timeslot] = timeslot_count
        #iterating over each slot of timetable
        for i, venue in zip(range(len(venues)),venues):
            for j, timeslot in zip(range(len(timeslots)),timeslots):
                if(sheet.cell(i+5,2).value is None): #if no venue, probably filler therefore continue
                continue
            slot=sheet.cell(i+5,j+3).value
            if(slot is None): #if slotvalue empty, it's an unoccupied slot
                slots.append(Slot(sheetnum+1,all_venues[venue],all_timeslots[timeslot],None,None,None))
            elif(len(slot.split('\n'))!=3): #if slotvalue cannot be divided into 3 parts, it's invalid therefore skip
                slots.append(Slot(sheetnum+1,all_venues[venue],all_timeslots[timeslot],None,None,None))
            else:
                slotvalues = slot.split('\n')
                #getting the course
                if(slotvalues[0].strip().upper() is not None):
                    courseObj=Course(slotvalues[0].strip().upper())
                    if courseObj not in all_courses:
                        course_count += 1
                        all_courses[courseObj] = course_count
                else:
                    courseObj=''

                #getting the teacher
                if(slotvalues[1].strip().upper() is not None):
                    teacherObj = Teacher(slotvalues[1].strip().upper())
                    if teacherObj not in all_teachers:
                        teacher_count += 1
                        all_teachers[teacherObj] = teacher_count
                else:
                    teacherObj=''

                #getting the section(also atomically dividing the attributes)
                sectionValues=slotvalues[2].replace('[','').replace(']','').replace('/','')
                if(sectionValues[0] in '12345678'):
                    sectionObj = Section(int(sectionValues[0]),(sectionValues[1:]))
                    if sectionObj not in all_sections:
                        section_count +=1
                        all_sections[sectionObj] = section_count
                else:
                    sectionObj=''

                slots.append(Slot(sheetnum+1,all_venues[venue],all_timeslots[timeslot],all_courses[courseObj],all_teachers[teacherObj],all_sections[sectionObj]))
if __name__ == "__main__":
    main()
