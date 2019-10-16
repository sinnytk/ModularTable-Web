from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook

STYLE_SLOT_NORMAL = NamedStyle('NORMAL')
STYLE_SLOT_VENUE = NamedStyle('VENUE')
STYLE_SLOT_TIMESLOT = NamedStyle('TIMESLOT')
STYLE_SLOT_NORMAL.alignment = STYLE_SLOT_VENUE.alignment = STYLE_SLOT_TIMESLOT.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
STYLE_SLOT_NORMAL.border = STYLE_SLOT_VENUE.border = STYLE_SLOT_TIMESLOT.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))




def generate_excel(slots=None, query_course_list=None, query_teacher_list=None, query_section_list=None):
    if not slots:
        return None
    day_list = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY"]
    wb = Workbook()
    wb.remove(wb['Sheet'])
    
    wb.add_named_style(STYLE_SLOT_NORMAL)
    wb.add_named_style(STYLE_SLOT_VENUE)
    wb.add_named_style(STYLE_SLOT_TIMESLOT)


    for daynum in range(len(day_list)):
        current_sheet = wb.create_sheet(day_list[daynum])
        
        # layout
        current_sheet.print_options.horizontalCentered = True
        current_sheet.print_options.verticalCentered = True
        current_sheet.page_setup.orientation = current_sheet.ORIENTATION_LANDSCAPE
        current_sheet.page_setup.paperSize = current_sheet.PAPERSIZE_LETTER
        current_sheet.page_setup.fitToPage = True
        current_sheet.page_setup.fitToWidth = 1
        current_sheet.page_setup.fitToHeight = 1
        # margins        
        current_sheet.page_margins.bottom = current_sheet.page_margins.left = current_sheet.page_margins.right = current_sheet.page_margins.header = 0.25
        current_sheet.page_margins.top = 0.5
        # headers
        current_sheet.oddHeader.center.text = "&[Tab]"
        current_sheet.oddHeader.center.size = 20
        current_sheet.oddHeader.center.font = "Arial"
        current_sheet.oddHeader.left.text = "ModularTable"
        current_sheet.oddHeader.left.size = 20
        current_sheet.oddHeader.left.font = "Arial"
        current_sheet.oddHeader.right.text = "Tarun Kumar"
        current_sheet.oddHeader.right.size = 20
        current_sheet.oddHeader.right.font = "Arial"
        
        current_slots = slots.filter(daynum=daynum+1)
        timeslot_list = current_slots.values('timeslot__starttime', 'timeslot__endtime').distinct().order_by('timeslot')
        venue_list = current_slots.values('venuenum','venuenum__venuename').distinct().order_by('venuenum')
        for row in current_sheet.iter_rows(min_col=2, max_col=len(timeslot_list)+1, max_row=1):
            for i, cell in enumerate(row):
                cell.style = 'TIMESLOT'
                cell.value = f"{timeslot_list[i]['timeslot__starttime']} - {timeslot_list[i]['timeslot__endtime']}"
        for row_count, row in enumerate(current_sheet.iter_rows(min_col=1, max_col=len(timeslot_list)+1, min_row=2, max_row=len(venue_list)+1)):
            row[0].style = 'VENUE'
            row[0].value = current_slots[row_count * (len(timeslot_list))].venuenum.venuename
            for cell_count, cell in enumerate(row[1:]):
                current_slot = current_slots[(row_count * (len(timeslot_list)))+cell_count]
                cell.style = 'NORMAL'
                if not current_slot.teachernum:
                    continue
                if (query_teacher_list is None or current_slot.teachernum.teachernum in query_teacher_list) and (query_course_list is None or current_slot.coursenum.coursenum in query_course_list) and (query_section_list is None or current_slot.sectionnum.sectionnum in query_section_list):
                    cell.value = f"{current_slot.teachernum.teachername}\n{current_slot.sectionnum.semester}{current_slot.sectionnum.section}\n{current_slot.coursenum.coursecode}"
        for i in range(2, len(timeslot_list)+2):
            current_sheet.column_dimensions[get_column_letter(i)].width = 24
        for i in range(2, len(venue_list)+2):
            current_sheet.row_dimensions[i].height = 60
            current_sheet.print_area = f'A1:{get_column_letter(len(timeslot_list)+1)}{len(venue_list)+1}'
        
    return save_virtual_workbook(wb)